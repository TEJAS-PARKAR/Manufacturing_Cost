from io import BytesIO

from openpyxl import Workbook

from backend.models import SupplierSessionResponse
from backend.services.negotiation_service import SupplierNegotiationService


def test_session_memory_resumes_supplier_context_across_sessions() -> None:
    service = SupplierNegotiationService()

    first_session = service.start_session(employee_id="EMP1001", part_number="123456789012")
    assert first_session["session_key"] == ("EMP1001", "123456789012")

    service.record_supplier_message(
        employee_id="EMP1001",
        part_number="123456789012",
        message="My part needs 1200 pieces, material is CRCA and coating is powder coating.",
    )

    resumed = service.get_session_context(employee_id="EMP1001", part_number="123456789012")

    assert resumed["employee_id"] == "EMP1001"
    assert resumed["part_number"] == "123456789012"
    assert len(resumed["history"]) >= 1
    assert resumed["summary"]


def test_excel_upload_extracts_costing_fields_into_session() -> None:
    service = SupplierNegotiationService()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Costing"
    worksheet.append(["quantity", "material", "material_rate", "coating", "process_information"])
    worksheet.append([1200, "CRCA", 65, "POWDER COATING", "LASER CUTTING"])

    buffer = BytesIO()
    workbook.save(buffer)
    payload = buffer.getvalue()

    result = service.ingest_excel(
        employee_id="EMP1001",
        part_number="123456789012",
        file_bytes=payload,
        filename="costing.xlsx",
    )

    assert result["extracted_data"]["quantity"] == 1200
    assert result["extracted_data"]["material"] == "CRCA"
    assert result["extracted_data"]["material_rate"] == 65.0
    assert result["extracted_data"]["coating"] == "POWDER COATING"


def test_excel_upload_builds_raw_table_and_interpretation_layers() -> None:
    service = SupplierNegotiationService()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Supplier Quote"
    worksheet.append(["Part Number", "Qty", "Material Grade", "Rate / Kg", "Surface Finish", "Process"])
    worksheet.append(["ABC-001", 1200, "CRCA", 65, "Powder Coating", "Laser Cutting"])

    buffer = BytesIO()
    workbook.save(buffer)
    payload = buffer.getvalue()

    result = service.ingest_excel(
        employee_id="EMP1001",
        part_number="123456789012",
        file_bytes=payload,
        filename="quote.xlsx",
    )

    assert "raw_table" not in result
    assert result["excel_interpretation"]["quantity"] == 1200
    assert result["excel_interpretation"]["material"] == "CRCA"
    assert result["excel_interpretation"]["material_rate"] == 65.0
    assert result["excel_interpretation"]["coating"] == "POWDER COATING"

    internal_session = service._ensure_session(
        employee_id="EMP1001",
        part_number="123456789012",
    )
    assert internal_session["raw_table"]["sheet_name"] == "Supplier Quote"
    assert internal_session["raw_table"]["headers"] == ["Part Number", "Qty", "Material Grade", "Rate / Kg", "Surface Finish", "Process"]


def test_wrapped_llm_extracted_data_is_preserved_in_response() -> None:
    service = SupplierNegotiationService()
    service._interpret_with_llm = lambda raw_table: {
        "extracted_data": {
            "quantity": 1200,
            "material": "CRCA",
            "material_rate": 65,
            "coating": "POWDER COATING",
        }
    }

    result = service._interpret_excel_table({"headers": [], "rows": []})

    assert result["quantity"] == 1200
    assert result["material"] == "CRCA"
    assert result["material_rate"] == 65.0
    assert result["coating"] == "POWDER COATING"


def test_empty_llm_interpretation_falls_back_to_headers() -> None:
    service = SupplierNegotiationService()
    service._interpret_with_llm = lambda raw_table: {
        "quantity": None,
        "material": None,
        "material_rate": None,
        "coating": None,
        "process_information": None,
    }

    result = service._interpret_excel_table({
        "headers": ["quantity", "material", "material_rate", "coating", "process_information"],
        "rows": [[1200, "CRCA", 65, "POWDER COATING", "LASER CUTTING"]],
    })

    assert result["quantity"] == 1200
    assert result["material"] == "CRCA"
    assert result["material_rate"] == 65.0
    assert result["coating"] == "POWDER COATING"


def test_session_response_preserves_allowance_gate_state() -> None:
    service = SupplierNegotiationService()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Costing"
    worksheet.append([
        "quantity",
        "material",
        "material_rate",
        "coating",
        "process_information",
        "part_length",
        "part_width",
        "part_thickness",
        "sheet_length",
        "sheet_width",
    ])
    worksheet.append([1200, "CRCA", 65, "POWDER COATING", "LASER CUTTING", 250, 200, 2, 1250, 2500])

    buffer = BytesIO()
    workbook.save(buffer)
    payload = buffer.getvalue()

    result = service.ingest_excel(
        employee_id="EMP1001",
        part_number="123456789012",
        file_bytes=payload,
        filename="costing.xlsx",
    )

    response = SupplierSessionResponse(**result)
    assert response.awaiting_allowance_response is True


def test_session_response_excludes_raw_excel_rows() -> None:
    service = SupplierNegotiationService()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Supplier Quote"
    worksheet.append(["Part Number", "Qty", "Material Grade", "Rate / Kg", "Surface Finish", "Process"])
    worksheet.append(["ABC-001", 1200, "CRCA", 65, "Powder Coating", "Laser Cutting"])

    buffer = BytesIO()
    workbook.save(buffer)
    payload = buffer.getvalue()

    result = service.ingest_excel(
        employee_id="EMP1001",
        part_number="123456789012",
        file_bytes=payload,
        filename="quote.xlsx",
    )

    public_response = SupplierSessionResponse(**result)
    public_payload = public_response.model_dump(exclude_none=True)

    assert "raw_table" not in public_payload
    assert public_payload["extracted_data"]["material"] == "CRCA"
