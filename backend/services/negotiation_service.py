from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
import json
import logging
import math
import os
import re
from typing import Any
import pandas as pd
import requests
from openpyxl import load_workbook
from backend.services.mongo_service import MongoConnection

logger = logging.getLogger(__name__)

class SupplierNegotiationService:
    REQUIRED_FIELDS = [
        "quantity",
        "dimensions",
        "material",
        "material_rate",
        "coating",
        "process_information",
    ]

    INTERNAL_BENCHMARKS = {
        "CRCA": {"material_rate": 65.0, "process_cost": 120.0, "overhead": 55.0},
        "ALUMINUM": {"material_rate": 220.0, "process_cost": 190.0, "overhead": 80.0},
        "STAINLESS": {"material_rate": 300.0, "process_cost": 250.0, "overhead": 100.0},
    }

    APPROVED_SHEET_SIZES = [
        (1250, 2500),
        (1500, 2500),
        (1250, 2700),
    ]

    def __init__(self) -> None:
        self.sessions: dict[tuple[str, str], dict[str, Any]] = {}
        self.groq_api_key = os.getenv("GROQ_API_KEY") or os.getenv("OPENAI_API_KEY")
        self.groq_model = os.getenv("GROQ_MODEL", "llama-3.1-8b-instant")
        collection_name = os.getenv("MONGODB_COLLECTION", "supplier_sessions")
        self.mongo_collection = MongoConnection.get_collection(collection_name)

    
    def start_session(
        self,
        employee_id: str,
        part_number: str
    ) -> dict[str, Any]:
        part_number = self._validate_part_number(part_number)
        key = self._session_key(
            employee_id,
            part_number
        )
        # Check memory first
        session = self.sessions.get(key)
        if session is not None:
            return self._serialize_session(session)
        # Check MongoDB before creating
        if self.mongo_collection is not None:
            doc = self.mongo_collection.find_one(
                {
                    "_id": self._storage_key(
                        employee_id,
                        part_number
                    )
                }
            )
            if doc is not None:
                session = self._hydrate_session(doc)
                self.sessions[key] = session
                logger.info("Existing session loaded from MongoDB for %s:%s", employee_id, part_number)
                return self._serialize_session(session)
        # Create new session only if nothing found
        session = {
            "employee_id": employee_id,
            "part_number": part_number,
            "session_key": key,
            "status": "active",
            "extracted_data": {},
            "raw_table": {},
            "excel_interpretation": {},
            "history": [],
            "summary": "New negotiation session started.",
            "missing_fields": self.REQUIRED_FIELDS,
            "review": {
                "recommendation": "review",
                "benchmark_reference": self._default_benchmark(
                    employee_id,
                    part_number
                ),
            },
            "negotiation": {
                "supplier_quote": 0,
                "predicted_cost": 0,
                "variance": 0,
                "ai_recommendation": "",
                "counter_offer": 0,
                "status": "pending",
                "rounds": [],
                "challenged_drivers": []
            }
        }
        self.sessions[key] = session
        self._persist_session(session)
        logger.info("New session created for %s:%s", employee_id, part_number)
        return self._serialize_session(session)

    
    def get_session_context(
        self,
        employee_id: str,
        part_number: str
    ) -> dict[str, Any]:
        session = self._ensure_session(
            employee_id,
            part_number
        )
        return self._serialize_session(session)


    def record_supplier_message(self, employee_id: str, part_number: str, message: str) -> dict[str, Any]:
        session = self._ensure_session(employee_id, part_number)
        parsed = self._extract_from_message(message)
        missing_fields = set(session["missing_fields"])
        allowed_updates = {}
        for key, value in parsed.items():
            if key in missing_fields:
                allowed_updates[key] = value
        session["extracted_data"].update(allowed_updates)
        session["history"].append(
            {
                "role": "supplier",
                "message": message,
                "timestamp": self._now_iso(),
            }
        )
        session["missing_fields"] = self._identify_missing_fields(
            session["extracted_data"]
        )
        session["summary"] = self._build_summary(session)
        session["review"]["recommendation"] = self._recommendation(
            session["extracted_data"]
        )
        self._persist_session(session)
        return self._serialize_session(session)

    
    def ingest_excel(self, employee_id, part_number, file_bytes, filename):
        session = self._ensure_session(
            employee_id,
            part_number
        )
        raw_table = self._extract_raw_table(file_bytes)
        logger.debug("Raw table extracted: %d rows, %d columns",
                      raw_table.get("row_count", 0), raw_table.get("column_count", 0))
        interpretation = self._interpret_excel_table(
            raw_table
        )
        logger.info("RAW TABLE HEADERS: %s", raw_table.get("headers"))
        logger.debug("INTERPRETATION: %s", interpretation)
        logger.debug("Excel interpretation result: %s", list(interpretation.keys()))
        session["raw_table"] = raw_table
        session["excel_interpretation"] = interpretation
        session.setdefault("revisions", [])
        for key, new_value in interpretation.items():
            old_value = session["extracted_data"].get(key)
            if old_value != new_value:
                session["revisions"].append({
                    "field": key,
                    "old_value": old_value,
                    "new_value": new_value,
                    "source": "excel_reupload",
                    "timestamp": self._now_iso()
                })
        session["extracted_data"].update(interpretation)
        session["missing_fields"] = self._identify_missing_fields(
            session["extracted_data"]
        )
        if session["missing_fields"]:
            session["history"].append(
                {
                    "role": "assistant",
                    "message": (
                        "The following mandatory fields are missing:\n\n• "
                        + "\n• ".join(session["missing_fields"])
                        + "\n\nPlease provide them through chat or upload a revised costing sheet."
                    ),
                    "timestamp": self._now_iso()
                }
            )
        logger.debug("Missing fields after recalculation: %s", session["missing_fields"])
        session["summary"] = self._build_summary(
            session
        )
        session["negotiation"] = (
            self.generate_negotiation_recommendation(
                session["extracted_data"]
            )
        )
        logger.error("########### INGEST EXCEL DEBUG HIT ###########")
        # Mark that the cutting allowance question needs to be answered
        session["awaiting_allowance_response"] = True
        # Clear previous sheet optimization so supplier must re-validate
        session["sheet_optimization"] = {}
        logger.debug("Extracted data keys after update: %s", list(session["extracted_data"].keys()))
        logger.info(
            "ALLOWANCE FLAG = %s",
            session.get("awaiting_allowance_response")
        )
        logger.info(
            "SERIALIZED FLAG = %s",
            self._serialize_session(session).get("awaiting_allowance_response")
        )
        self._persist_session(session)
        return self._serialize_session(session)

    
    def generate_negotiation_recommendation(
        self,
        extracted_data
    ):
        supplier_quote = round(float(
            extracted_data.get("total_cost", 0)
        ), 2)
        expected_cost = round(self._compute_expected_cost(extracted_data), 2)
        variance = 0
        if expected_cost > 0:
            variance = round(
                (
                    (supplier_quote - expected_cost)
                    / expected_cost
                ) * 100,
                2
            )
        if variance <= 5:
            recommendation = "approve"
            counter_offer = supplier_quote
        elif variance <= 15:
            recommendation = "negotiate"
            counter_offer = round(
                expected_cost * 1.03,
                2
            )
        else:
            recommendation = "reject"
            counter_offer = expected_cost
        drivers = self._rank_negotiation_drivers(extracted_data)[:3]
        return {
            "supplier_quote": supplier_quote,
            "predicted_cost": expected_cost,
            "variance": variance,
            "ai_recommendation": recommendation,
            "counter_offer": counter_offer,
            "negotiation_drivers": drivers,
            "status": "active",
            "rounds": [],
            "challenged_drivers": []
        }

    def submit_for_review(self, employee_id: str, part_number: str) -> dict[str, Any]:
        session = self._ensure_session(employee_id, part_number)
        session["status"] = "submitted_for_review"
        session["history"].append(
            {
                "role": "system",
                "message": "Supplier session submitted to Tata Motors review dashboard.",
                "timestamp": self._now_iso(),
            }
        )
        session["summary"] = self._build_summary(session)
        self._persist_session(session)
        logger.info("Session submitted for review: %s:%s", employee_id, part_number)
        return self._serialize_session(session)

    def get_review_dashboard(self, employee_id: str, part_number: str) -> dict[str, Any]:
        session = self._ensure_session(employee_id, part_number)
        return {
            "session": self._serialize_session(session),
            "recommendation": self._recommendation(session["extracted_data"]),
            "benchmark_comparison": self._benchmark_comparison(session["extracted_data"]),
        }

    def approve_cost_inputs(self, employee_id: str, part_number: str, approval_payload: dict[str, Any]) -> dict[str, Any]:
        session = self._ensure_session(employee_id, part_number)
        approved = approval_payload.get("approved_values", {})
        session["extracted_data"].update(approved)
        session["status"] = "approved"
        session["history"].append(
            {
                "role": "tata",
                "message": "Approved cost inputs updated into final validated estimate.",
                "timestamp": self._now_iso(),
            }
        )
        session["summary"] = self._build_summary(session)
        self._persist_session(session)
        logger.info("Session approved: %s:%s", employee_id, part_number)
        return self._serialize_session(session)

    
    def _ensure_session(self, employee_id: str, part_number: str) -> dict[str, Any]:
        part_number = self._validate_part_number(part_number)
        key = self._session_key(employee_id, part_number)
        storage_key = self._storage_key(employee_id, part_number)
        session = self.sessions.get(key)
        # Memory cache exists
        if session is not None:
            if self.mongo_collection is not None:
                doc = self.mongo_collection.find_one(
                    {"_id": storage_key}
                )
                # Deleted from Mongo
                if doc is None:
                    self.sessions.pop(key, None)
                    return self.start_session(
                        employee_id,
                        part_number
                    )
            return session
        # Not in memory -> load from Mongo
        if self.mongo_collection is not None:
            doc = self.mongo_collection.find_one(
                {"_id": storage_key}
            )
            if doc:
                session = self._hydrate_session(doc)
                self.sessions[key] = session
                return session
        return self.start_session(
            employee_id,
            part_number
        )


    def _serialize_session(self, session: dict[str, Any]) -> dict[str, Any]:
        logger.info(
            "SERIALIZE FLAG = %s",
            session.get("awaiting_allowance_response")
        )
        return {
            "employee_id": session["employee_id"],
            "part_number": session["part_number"],
            "session_key": session["session_key"],
            "status": session["status"],
            "extracted_data": session["extracted_data"],
            "raw_table": session.get("raw_table", {}),
            "excel_interpretation": session.get("excel_interpretation", {}),
            "history": session["history"],
            "summary": session["summary"],
            "missing_fields": session["missing_fields"],
            "review": session["review"],
            "revisions": session.get("revisions", []),
            "negotiation": session.get(
                "negotiation",
                {}
            ),
            "sheet_optimization": session.get(
                "sheet_optimization",
                {}
            ),
            "awaiting_allowance_response": session.get(
                "awaiting_allowance_response",
                False
            ),
            "rejection_remark": session.get(
                "rejection_remark",
                None
            ),
        }

    def _storage_key(self, employee_id: str, part_number: str) -> str:
        return f"{employee_id.strip()}::{part_number.strip()}"

    
    def _persist_session(self, session: dict[str, Any]) -> None:
        if self.mongo_collection is None:
            logger.debug("No MongoDB collection available, skipping persistence")
            return
        doc = self._serialize_session(session)
        doc.pop("raw_table", None)
        doc["_id"] = self._storage_key(
            session["employee_id"],
            session["part_number"]
        )
        result = self.mongo_collection.replace_one(
            {"_id": doc["_id"]},
            doc,
            upsert=True
        )
        logger.debug("Session persisted: id=%s, matched=%d, modified=%d, upserted=%s",
                      doc["_id"], result.matched_count, result.modified_count, result.upserted_id)


    def _hydrate_session(self, document: dict[str, Any]) -> dict[str, Any]:
        session = dict(document)
        session.pop("_id", None)
        session_key = session.get("session_key")
        if isinstance(session_key, list):
            session["session_key"] = tuple(session_key)
        session["missing_fields"] = self._identify_missing_fields(
            session.get("extracted_data", {})
        )
        session.setdefault("sheet_optimization", {})

        # Preserve explicit persisted state on resume to avoid dropping the allowance gate.
        if "awaiting_allowance_response" not in session:
            extracted = session.get("extracted_data", {})
            dimension_keys = (
                "part_length",
                "length",
                "part_width",
                "width",
                "part_thickness",
                "thickness",
            )
            has_dimensions = any(key in extracted for key in dimension_keys)
            session["awaiting_allowance_response"] = bool(
                extracted and has_dimensions and not session.get("sheet_optimization")
            )
        else:
            session["awaiting_allowance_response"] = bool(
                session.get("awaiting_allowance_response", False)
            )
        return session

    def _validate_part_number(self, part_number: str) -> str:
        pn = str(part_number).strip()
        if not pn.isdigit() or len(pn) != 12:
            raise ValueError("Part number must be exactly 12 digits (0-9 only).")
        return pn

    def _session_key(self, employee_id: str, part_number: str) -> tuple[str, str]:
        return (employee_id.strip(), part_number.strip())

    def _now_iso(self) -> str:
        return datetime.now(timezone.utc).isoformat()

    def _build_summary(self, session: dict[str, Any]) -> str:
        data = session["extracted_data"]
        material = data.get("material") or "pending"
        quantity = data.get("quantity") or "pending"
        dimensions = data.get("dimensions") or "pending"
        coating = data.get("coating") or "pending"
        return (
            f"Supplier {session['employee_id']} is discussing part {session['part_number']}. "
            f"Current extracted context: quantity={quantity}, material={material}, dimensions={dimensions}, coating={coating}."
        )

    def _identify_missing_fields(self, extracted_data: dict[str, Any]) -> list[str]:
        missing = []
        if not extracted_data.get("quantity"):
            missing.append("quantity")
        if not extracted_data.get("dimensions"):
            missing.append("dimensions")
        if (not extracted_data.get("material") and not extracted_data.get("material_grade")):
            missing.append("material")
        if not extracted_data.get("material_rate"):
            missing.append("material_rate")
        if not extracted_data.get("coating") and not extracted_data.get("coating_cost"):
            missing.append("coating")
        if not extracted_data.get("process_information"):
            missing.append("process_information")
        return missing

    def _recommendation(self, extracted_data: dict[str, Any]) -> str:
        material_rate = float(extracted_data.get("material_rate") or 0)
        benchmark = self.INTERNAL_BENCHMARKS.get((extracted_data.get("material") or "").upper(), {}).get("material_rate", 0.0)
        if benchmark and material_rate > 0:
            if material_rate <= benchmark * 1.05:
                return "accept"
            if material_rate <= benchmark * 1.15:
                return "review"
            return "negotiate_further"
        return "review"

    def _default_benchmark(self, employee_id: str, part_number: str) -> dict[str, Any]:
        return {
            "employee_id": employee_id,
            "part_number": part_number,
            "raw_material_rate": 65.0,
            "process_cost": 120.0,
            "coating_cost": 40.0,
            "benchmark_note": "Internal Tata Motors benchmark reference loaded for review.",
        }

    def _extract_from_message(self, message: str) -> dict[str, Any]:
        normalized = message.lower()
        extracted: dict[str, Any] = {}
        quantity_match = re.search(r"(\d+(?:,\d+)*)\s*(pieces?|pcs?|units?|qty)", normalized)
        if quantity_match:
            extracted["quantity"] = int(quantity_match.group(1).replace(",", ""))
        dimension_match = re.search(r"(\d+(?:\.\d+)?)\s*(?:x|by|×)\s*(\d+(?:\.\d+)?)\s*(?:x|by|×)\s*(\d+(?:\.\d+)?)", normalized)
        if dimension_match:
            extracted["dimensions"] = [
                float(dimension_match.group(1)),
                float(dimension_match.group(2)),
                float(dimension_match.group(3)),
            ]
        material_map = {
            "crca": "CRCA",
            "mild steel": "MILD STEEL",
            "stainless": "STAINLESS",
            "aluminum": "ALUMINUM",
        }
        for keyword, value in material_map.items():
            if keyword in normalized:
                extracted["material"] = value
                break
        coating_keywords = {
            "powder coating": "POWDER COATING",
            "painting": "PAINTING",
            "zinc": "ZINC COATING",
            "chrome": "CHROME COATING",
        }
        for keyword, value in coating_keywords.items():
            if keyword in normalized:
                extracted["coating"] = value
                break
        process_keywords = []
        if "laser" in normalized or "cut" in normalized:
            process_keywords.append("LASER CUTTING")
        if "bend" in normalized:
            process_keywords.append("BENDING")
        if "weld" in normalized:
            process_keywords.append("WELDING")
        if "drill" in normalized or "hole" in normalized:
            process_keywords.append("DRILLING")
        if process_keywords:
            extracted["process_information"] = process_keywords
        rate_match = re.search(r"material rate\s*(?:is|=|:)?\s*(\d+(?:\.\d+)?)", normalized)
        if rate_match:
            extracted["material_rate"] = float(rate_match.group(1))
        return extracted

    def _extract_raw_table(self, file_bytes: bytes) -> dict[str, Any]:
        buffer = BytesIO(file_bytes)
        workbook = load_workbook(buffer, data_only=True)
        worksheet = workbook.active
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        headers = [self._clean_cell(value) for value in rows[0]] if rows else []
        data_rows = []
        for row in rows[1:]:
            data_rows.append([self._clean_cell(value) for value in row])
        dataframe = pd.DataFrame(data_rows, columns=headers) if headers else pd.DataFrame(data_rows)
        return {
            "sheet_name": worksheet.title,
            "headers": headers,
            "rows": data_rows,
            "row_count": len(data_rows),
            "column_count": len(headers),
            "dataframe_preview": dataframe.to_dict(orient="records"),
        }


    def _extract_dimensions_from_raw_table(self, rows):
        for row in rows:
            text = " ".join(
                str(x)
                for x in row
                if x is not None
            ).upper()
            if (
                "TH" in text
                and "WD" in text
                and "LG" in text
            ):
                nums = []
                for item in row:
                    try:
                        nums.append(float(item))
                    except (ValueError, TypeError):
                        pass
                if len(nums) >= 3:
                    return {
                        "thickness": nums[0],
                        "width": nums[1],
                        "length": nums[2],
                        "dimensions": nums[:3]
                    }
        return {}


    def _interpret_excel_table(self, raw_table: dict[str, Any]) -> dict[str, Any]:
        llm_result = self._interpret_with_llm(raw_table) or {}
        dimensions = self._extract_dimensions_from_raw_table(
            raw_table.get("rows", [])
        )
        llm_result.update(dimensions)
        if llm_result:
            return self._normalize_interpreted_values(llm_result)
        headers = [self._normalize_header(header) for header in raw_table.get("headers", [])]
        rows = raw_table.get("rows", [])
        if not headers or not rows:
            return {}
        first_row = rows[0]
        lookup: dict[str, Any] = {}
        for index, header in enumerate(headers):
            lookup[self._normalize_header_key(header)] = first_row[index] if index < len(first_row) else None
        interpreted: dict[str, Any] = {}
        quantity = self._extract_quantity_from_lookup(lookup)
        if quantity is not None:
            interpreted["quantity"] = quantity
        dimensions = self._extract_dimensions_from_lookup(lookup)
        if dimensions:
            interpreted["dimensions"] = dimensions
        material = self._extract_material_from_lookup(lookup)
        if material:
            interpreted["material"] = material
        material_rate = self._extract_numeric_from_lookup(lookup, ["material_rate", "rate_per_kg", "rate", "rate_kg"])
        if material_rate is not None:
            interpreted["material_rate"] = material_rate
        coating = self._extract_coating_from_lookup(lookup)
        if coating:
            interpreted["coating"] = coating
        process_information = self._extract_process_information_from_lookup(lookup)
        if process_information:
            interpreted["process_information"] = process_information
        return interpreted
    
    def _interpret_with_llm(self, raw_table: dict[str, Any]) -> dict[str, Any]:
        if not self.groq_api_key:
            return {}
        try:
            # H2: Truncate rows to avoid context window overflow on large sheets
            rows_to_send = raw_table.get("rows", [])[:50]
            row_note = ""
            if len(raw_table.get("rows", [])) > 50:
                row_note = f" (showing first 50 of {len(raw_table['rows'])} rows)"

            payload = {
                "model": self.groq_model,
                "messages": [
                    {
                        "role": "system",
                        "content": """You are an expert in Tata Motors supplier costing sheets.

Analyze the costing sheet and return a FLAT JSON object with SCALAR values only.
Do NOT return arrays or nested objects for the top-level fields (except process_information).

RETURN THESE EXACT KEYS (use null if not found):

  "quantity"            : integer — production batch quantity (e.g. 132)
  "material"            : string — single material name/grade (e.g. "10 MM E 46", "CRCA")
  "material_grade"      : string — grade only (e.g. "E 46")
  "material_rate"       : number — Rs/kg rate from the raw material section (e.g. 60.3)

  SHEET DIMENSIONS (from "Full Sheet Size" row):
  "sheet_length"        : number — Full Sheet length in mm
  "sheet_width"         : number — Full Sheet width in mm
  "sheet_thickness"     : number — Full Sheet thickness in mm

  PART DIMENSIONS (from "Shear Size" or "Blank Size" row):
  "part_length"         : number — Shear/Blank length in mm
  "part_width"          : number — Shear/Blank width in mm
  "part_thickness"      : number — Shear/Blank thickness in mm

  BACKWARD COMPAT (populate from sheet/part fields):
  "thickness"           : number — same as sheet_thickness
  "width"               : number — same as part_width
  "length"              : number — same as part_length

  "gross_weight"        : number — gross weight per piece in kg
  "finished_weight"     : number — finished weight per piece in kg (e.g. 1.25)
  "scrap_weight"        : number — scrap weight per piece in kg
  "blank_weight"        : number — blank weight per piece in kg
  "raw_material_cost"   : number — NET material cost per piece (look for "NET MATL. COST" or "RM COST")
  "conversion_cost"     : number — total conversion cost per piece (look for "TOTAL CON. COST")
  "coating_cost"        : number — sum of all coating/surface protection costs per piece
  "overhead_cost"       : number — overhead per piece
  "icc_cost"            : number — ICC on raw material per piece
  "rejection_cost"      : number — rejection allowance per piece
  "profit"              : number — profit per piece
  "packing_cost"        : number — packing cost per piece
  "transport_cost"      : number — transport cost per piece
  "total_cost"          : number — final TOTAL cost per piece (the last/bottom "TOTAL" in the sheet)
  "coating"             : string — coating type (e.g. "POWDER COATING", "ZINC PLATING")
  "process_information" : array of {"process": string, "cost": number} — individual process line items

CRITICAL RULES:
- "material" must be a SINGLE STRING, not a list or array.
- Every numeric field must be a single number, never an expression.
- Extract "Full Sheet Size" dimensions into sheet_length, sheet_width, sheet_thickness.
- Extract "Shear Size" or "Blank Size" dimensions into part_length, part_width, part_thickness.
- If only one set of dimensions exists, use it for both sheet_* and part_* fields.
- Also copy sheet_thickness to thickness, part_width to width, part_length to length.
- Sum all coating-related line items (powder coating + shot blasting + primer) into coating_cost.
- Use the "NET MATL. COST PER PIECE" row value as raw_material_cost.
- Use the final "TOTAL" row at the bottom of the cost summary as total_cost.
- Use null if a field is genuinely not present in the sheet.
"""
                    },
                    {
                        "role": "user",
                        "content": json.dumps(
                            {
                                "sheet_name": raw_table.get("sheet_name"),
                                "headers": raw_table.get("headers"),
                                "rows": rows_to_send,
                                "note": row_note,
                            },
                            ensure_ascii=False
                        )
                    }
                ],
                "temperature": 0.1,
                "response_format": {"type": "json_object"},
            }
            logger.debug(
                "Sending to Groq: %d rows (of %d total), %d columns",
                len(rows_to_send),
                len(raw_table.get("rows", [])),
                len(raw_table.get("headers", []))
            )
            response = requests.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {self.groq_api_key}",
                    "Content-Type": "application/json"
                },
                json=payload,
                timeout=30
            )
            if response.status_code != 200:
                logger.error(
                    "Groq Error %s: %s",
                    response.status_code,
                    response.text
                )
            response.raise_for_status()
            logger.debug(
                "Groq API responded with status %d",
                response.status_code
            )
            content = response.json()["choices"][0]["message"]["content"]
            content = content.strip()
            # Fix things like:
            # "coating_cost": 3.25 + 0.89
            content = re.sub(
                r'(:\s*)(\d+(?:\.\d+)?)\s*\+\s*(\d+(?:\.\d+)?)',
                lambda m: f"{m.group(1)}{float(m.group(2)) + float(m.group(3))}",
                content
            )
            if content.startswith("```"):
                content = re.sub(r"^```json", "", content)
                content = content.replace("```", "")
                content = content.strip()

            parsed = None
            try:
                parsed = json.loads(content)
                logger.info("PARSED DATA: %s", parsed)
            except json.JSONDecodeError as e:
                # Fallback: try to extract the first JSON object from the response
                logger.warning("Direct JSON parse failed: %s — attempting regex extraction", str(e))
                json_match = re.search(r'\{[\s\S]*\}', content)
                if json_match:
                    try:
                        parsed = json.loads(json_match.group())
                        logger.info("PARSED DATA (regex fallback): %s", parsed)
                    except json.JSONDecodeError:
                        logger.error("Regex fallback also failed. RAW CONTENT:\n%s", content)
                        return {}
                else:
                    logger.error("No JSON object found in LLM response. RAW CONTENT:\n%s", content)
                    return {}

            logger.debug(
                "Groq output keys: %s",
                list(parsed.keys()) if isinstance(parsed, dict)
                else type(parsed)
            )
            if isinstance(parsed, dict):
                return parsed
            return {}

        except Exception as e:
            logger.warning(
                "Groq LLM interpretation failed: %s",
                str(e)
            )
            return {}

    def _normalize_interpreted_values(self, values: dict[str, Any]) -> dict[str, Any]:
        normalized: dict[str, Any] = {}
        if values.get("quantity") is not None:
            try:
                normalized["quantity"] = int(float(values["quantity"]))
            except (ValueError, TypeError):
                pass
        if values.get("dimensions"):
            if isinstance(values["dimensions"], list):
                normalized["dimensions"] = [float(item) for item in values["dimensions"]]
        if (
            values.get("thickness") is not None
            and values.get("width") is not None
            and values.get("length") is not None
        ):
            try:
                normalized["dimensions"] = [
                    round(float(values["thickness"]), 2),
                    round(float(values["width"]), 2),
                    round(float(values["length"]), 2),
                ]
            except (ValueError, TypeError):
                pass

        # --- Handle separate sheet vs part dimensions ---
        for prefix in ("sheet", "part"):
            for dim in ("length", "width", "thickness"):
                key = f"{prefix}_{dim}"
                val = values.get(key)
                if val is not None:
                    try:
                        normalized[key] = round(float(val), 2)
                    except (ValueError, TypeError):
                        pass

        # Populate part_* from old fields if not already set
        if not normalized.get("part_length") and normalized.get("length"):
            normalized["part_length"] = round(float(normalized["length"]), 2)
        if not normalized.get("part_width") and normalized.get("width"):
            normalized["part_width"] = round(float(normalized["width"]), 2)
        if not normalized.get("part_thickness") and normalized.get("thickness"):
            normalized["part_thickness"] = round(float(normalized["thickness"]), 2)

        # Populate sheet_* from old fields if not already set
        if not normalized.get("sheet_length"):
            normalized["sheet_length"] = normalized.get("part_length", 0)
        if not normalized.get("sheet_width"):
            normalized["sheet_width"] = normalized.get("part_width", 0)
        if not normalized.get("sheet_thickness"):
            normalized["sheet_thickness"] = normalized.get("part_thickness", 0)

        # Rebuild backward-compat dimensions from part fields
        if normalized.get("part_thickness") and normalized.get("part_width") and normalized.get("part_length"):
            normalized["dimensions"] = [
                normalized["part_thickness"],
                normalized["part_width"],
                normalized["part_length"],
            ]

        # --- Handle material (may be string, list-of-dicts, or stringified list) ---
        raw_material = values.get("material")
        material = self._extract_material_name(raw_material)
        if material:
            normalized["material"] = material

        if values.get("material_rate") is not None:
            try:
                normalized["material_rate"] = float(values["material_rate"])
            except (ValueError, TypeError):
                pass

        coating = self._normalize_coating(values.get("coating"))
        if coating:
            normalized["coating"] = coating

        process_information = values.get("process_information")
        if process_information and isinstance(process_information, list):
            processes = []
            for item in process_information:
                if isinstance(item, dict):
                    processes.append(
                        {
                            "process": item.get("process"),
                            "cost": item.get("cost", 0)
                        }
                    )
                elif isinstance(item, str):
                    processes.append(
                        {
                            "process": self._normalize_process(item),
                            "cost": 0
                        }
                    )
            normalized["process_information"] = processes

        # Additional fields extracted by Groq (all rounded to 2 dp)
        for field in [
            "material_grade", "thickness", "width", "length",
            "finished_weight", "scrap_weight", "blank_weight",
            "gross_weight",
            "raw_material_cost", "conversion_cost", "coating_cost",
            "overhead_cost", "icc_cost", "rejection_cost",
            "profit", "packing_cost", "transport_cost",
            "total_cost",
        ]:
            val = values.get(field)
            if val is not None:
                try:
                    normalized[field] = round(float(val), 2)
                except (ValueError, TypeError):
                    normalized[field] = val

        if normalized.get("coating_cost") and not normalized.get("coating"):
            normalized["coating"] = "SURFACE PROTECTION"

        # --- Post-process: fill missing cost fields from process_information ---
        self._postprocess_from_process_info(normalized)

        # --- If material_rate still missing, try to extract from material list ---
        if not normalized.get("material_rate") and isinstance(raw_material, list):
            for entry in raw_material:
                if isinstance(entry, dict):
                    rate = entry.get("MATERIAL_RATE") or entry.get("material_rate")
                    if rate and float(rate) > 0:
                        normalized["material_rate"] = float(rate)
                        break

        # --- If quantity still missing, try to extract from material list ---
        if not normalized.get("quantity") and isinstance(raw_material, list):
            for entry in raw_material:
                if isinstance(entry, dict):
                    qty = entry.get("QUANTITY") or entry.get("quantity")
                    if qty is not None:
                        try:
                            q = int(float(qty))
                            if q > 0:
                                normalized["quantity"] = q
                                break
                        except (ValueError, TypeError):
                            pass

        return normalized

    def _extract_material_name(self, raw_material: Any) -> str | None:
        """Extract a single material name from various LLM output formats."""
        if raw_material is None:
            return None

        # Simple string: "CRCA" or "10 MM E 46"
        if isinstance(raw_material, str):
            stripped = raw_material.strip()
            # Check if it looks like a stringified Python list: "[{...}]"
            if stripped.startswith("[") and "MATERIAL" in stripped.upper():
                # Try to extract the material name via regex
                mat_match = re.search(
                    r"['\"]MATERIAL['\"]\s*:\s*['\"]([^'\"]+)['\"]",
                    stripped, re.IGNORECASE
                )
                if mat_match:
                    return self._normalize_material(mat_match.group(1))
            return self._normalize_material(stripped)

        # List of dicts: [{"MATERIAL": "10 MM E 46", ...}, ...]
        if isinstance(raw_material, list) and raw_material:
            first = raw_material[0]
            if isinstance(first, dict):
                name = first.get("MATERIAL") or first.get("material") or first.get("type")
                if name:
                    return self._normalize_material(str(name))
            if isinstance(first, str):
                return self._normalize_material(first)

        return self._normalize_material(raw_material)

    def _postprocess_from_process_info(self, normalized: dict[str, Any]) -> None:
        """Fill missing cost component fields by matching process_information line items."""
        processes = normalized.get("process_information")
        if not processes or not isinstance(processes, list):
            return

        # Mapping: regex pattern on process name -> target field name
        COST_FIELD_PATTERNS = [
            (r"NET\s*MATL|R\.?\s*M\.?\s*COST|RAW\s*MATERIAL\s*COST", "raw_material_cost"),
            (r"TOTAL\s*CON\.?\s*COST|CONVERSION\s*COST", "conversion_cost"),
            (r"OVERHEAD", "overhead_cost"),
            (r"I\.?C\.?C", "icc_cost"),
            (r"REJECTION\b(?!.*RECOVERY)", "rejection_cost"),
            (r"PROFIT", "profit"),
            (r"PACKING", "packing_cost"),
            (r"TRANSPORT", "transport_cost"),
        ]
        COATING_PATTERNS = [
            r"POWDER\s*COAT", r"SHOT\s*BLAST", r"PRIMER",
            r"SURFACE\s*PROTECTION", r"ZINC\s*PLAT", r"PAINT",
        ]

        coating_sum = 0.0
        has_coating = False

        for item in processes:
            if not isinstance(item, dict):
                continue
            name = str(item.get("process") or "").upper()
            cost = item.get("cost", 0)
            try:
                cost = float(cost)
            except (ValueError, TypeError):
                continue

            # Check coating patterns first
            for pattern in COATING_PATTERNS:
                if re.search(pattern, name, re.IGNORECASE):
                    if cost > 0:
                        coating_sum += cost
                        has_coating = True
                    # Also set coating type if not set
                    if not normalized.get("coating"):
                        if "POWDER" in name:
                            normalized["coating"] = "POWDER COATING"
                        elif "ZINC" in name:
                            normalized["coating"] = "ZINC PLATING"
                        elif "PAINT" in name:
                            normalized["coating"] = "PAINTING"
                        else:
                            normalized["coating"] = "SURFACE PROTECTION"
                    break

            # Check cost field patterns
            for pattern, field in COST_FIELD_PATTERNS:
                if re.search(pattern, name, re.IGNORECASE):
                    if not normalized.get(field) and cost > 0:
                        normalized[field] = cost
                    break

        # Set coating_cost from summed coating items
        if has_coating and not normalized.get("coating_cost"):
            normalized["coating_cost"] = round(coating_sum, 2)

        # Extract total_cost from the last "TOTAL" entry
        if not normalized.get("total_cost"):
            for item in reversed(processes):
                if not isinstance(item, dict):
                    continue
                name = str(item.get("process") or "").upper().strip()
                if name == "TOTAL" or name.startswith("TOTAL"):
                    cost = item.get("cost", 0)
                    try:
                        cost = float(cost)
                        if cost > 0:
                            normalized["total_cost"] = cost
                            break
                    except (ValueError, TypeError):
                        continue

    def _normalize_header(self, header: Any) -> str:
        return str(header).strip() if header is not None else ""

    def _normalize_header_key(self, header: str) -> str:
        normalized = re.sub(r"[^a-z0-9]+", "_", header.strip().lower())
        return normalized.strip("_")

    def _clean_cell(self, value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return value

    def _extract_quantity_from_lookup(self, lookup: dict[str, Any]) -> int | None:
        for key in ["quantity", "qty", "qty_pcs", "pieces", "piece_count"]:
            value = lookup.get(key)
            if value is not None and str(value).strip() != "":
                try:
                    return int(float(str(value)))
                except ValueError:
                    pass
        for value in lookup.values():
            if isinstance(value, (int, float)):
                return int(float(value))
        return None

    def _extract_dimensions_from_lookup(self, lookup: dict[str, Any]) -> list[float] | None:
        dimension_values = []
        for key in ["length", "width", "height", "thickness"]:
            value = lookup.get(key)
            if value is not None and str(value).strip() != "":
                try:
                    dimension_values.append(float(value))
                except (TypeError, ValueError):
                    continue
        if dimension_values:
            return dimension_values[:3]
        return None

    def _extract_numeric_from_lookup(self, lookup: dict[str, Any], keys: list[str]) -> float | None:
        for key in keys:
            value = lookup.get(key)
            if value is not None and str(value).strip() != "":
                try:
                    return float(value)
                except (TypeError, ValueError):
                    continue
        return None

    def _extract_material_from_lookup(self, lookup: dict[str, Any]) -> str | None:
        for key in ["material", "material_grade", "material_type"]:
            value = lookup.get(key)
            if value is not None and str(value).strip() != "":
                return self._normalize_material(value)
        return None

    def _extract_coating_from_lookup(self, lookup: dict[str, Any]) -> str | None:
        for key in ["coating", "surface_coating", "surface_finish", "finish"]:
            value = lookup.get(key)
            if value is not None and str(value).strip() != "":
                return self._normalize_coating(value)
        return None

    def _extract_process_information_from_lookup(self, lookup: dict[str, Any]) -> list[str] | None:
        for key in ["process_information", "process", "processes"]:
            value = lookup.get(key)
            if value is not None and str(value).strip() != "":
                if isinstance(value, list):
                    normalized = [self._normalize_process(item) for item in value if self._normalize_process(item)]
                    if normalized:
                        return normalized
                normalized = self._normalize_process(value)
                if normalized:
                    return [normalized]
        return None

    def _normalize_material(self, value: Any) -> str | None:
        if value is None:
            return None
        normalized = str(value).strip().upper()
        mapping = {
            "CRCA": "CRCA",
            "MILD STEEL": "MILD STEEL",
            "MILDSTEEL": "MILD STEEL",
            "STAINLESS": "STAINLESS",
            "ALUMINUM": "ALUMINUM",
        }
        if normalized in mapping:
            return mapping[normalized]
        if "CRCA" in normalized:
            return "CRCA"
        if "MILD STEEL" in normalized or "MILDSTEEL" in normalized:
            return "MILD STEEL"
        if "STAINLESS" in normalized:
            return "STAINLESS"
        if "ALUMINUM" in normalized:
            return "ALUMINUM"
        return normalized or None

    def _normalize_coating(self, value: Any) -> str | None:
        if value is None:
            return None
        normalized = str(value).strip().upper()
        mapping = {
            "POWDER COATING": "POWDER COATING",
            "POWDER": "POWDER COATING",
            "PAINTING": "PAINTING",
            "ZINC": "ZINC COATING",
            "ZINC COATING": "ZINC COATING",
            "CHROME": "CHROME COATING",
            "CHROME COATING": "CHROME COATING",
        }
        if normalized in mapping:
            return mapping[normalized]
        if "POWDER" in normalized:
            return "POWDER COATING"
        if "PAINT" in normalized:
            return "PAINTING"
        if "ZINC" in normalized:
            return "ZINC COATING"
        if "CHROME" in normalized:
            return "CHROME COATING"
        return normalized or None

    def _normalize_process(self, value: Any) -> str | None:
        if value is None:
            return None
        normalized = str(value).strip().upper()
        if not normalized:
            return None
        replacements = {
            "LASER CUT": "LASER CUTTING",
            "LASER CUTTING": "LASER CUTTING",
            "CUTTING": "LASER CUTTING",
            "BEND": "BENDING",
            "BENDING": "BENDING",
            "WELD": "WELDING",
            "WELDING": "WELDING",
            "DRILL": "DRILLING",
            "DRILLING": "DRILLING",
            "HOLE": "DRILLING",
        }
        for source, target in replacements.items():
            if source in normalized:
                return target
        return normalized

    def _benchmark_comparison(self, extracted_data: dict[str, Any]) -> dict[str, Any]:
        material = (extracted_data.get("material") or "CRCA").upper()
        rate = float(extracted_data.get("material_rate") or 0)
        benchmark = self.INTERNAL_BENCHMARKS.get(material, {"material_rate": 65.0, "process_cost": 120.0})
        benchmark_rate = float(benchmark.get("material_rate", 0.0))
        variance = round(rate - benchmark_rate, 2) if benchmark_rate else 0.0
        return {
            "supplier_material_rate": rate,
            "internal_benchmark_rate": benchmark_rate,
            "variance": variance,
            "recommendation": self._recommendation(extracted_data),
        }

    # ─── Sheet Utilization Optimization ───────────────────────────────────
    def _validate_sheet_optimization(
        self,
        extracted_data: dict[str, Any],
        includes_cutting_allowance: bool = True,
    ) -> dict[str, Any]:
        """
        Calculate sheet utilization for all approved sheet sizes.

        Selection criteria:
        Lowest Weight Per Part =
        (Sheet Length × Sheet Width × Thickness × 7.85 / 10^6)
        / Number Of Parts Fitting

        This follows the procurement requirement provided by Tata Motors.
        """
        part_length = float(
            extracted_data.get("part_length")
            or extracted_data.get("length")
            or 0
        )
        part_width = float(
            extracted_data.get("part_width")
            or extracted_data.get("width")
            or 0
        )
        thickness = float(
            extracted_data.get("part_thickness")
            or extracted_data.get("sheet_thickness")
            or extracted_data.get("thickness")
            or 0
        )
        if part_length <= 0 or part_width <= 0 or thickness <= 0:
            return {
                "optimal": None,
                "message": "Part dimensions or thickness missing.",
                "all_options": [],
            }
        # Apply cutting allowance only if not already included
        if not includes_cutting_allowance:
            effective_length = part_length + (1.5 * thickness)
            effective_width = part_width + (1.5 * thickness)
        else:
            effective_length = part_length
            effective_width = part_width
        all_options = []
        best_weight_per_part = float("inf")
        best_option = None
        for sheet_l, sheet_w in self.APPROVED_SHEET_SIZES:
            # Normal orientation
            pieces_normal = (
                math.floor(sheet_l / effective_length)
                * math.floor(sheet_w / effective_width)
            )
            # Rotated orientation
            pieces_rotated = (
                math.floor(sheet_l / effective_width)
                * math.floor(sheet_w / effective_length)
            )
            pieces = max(
                pieces_normal,
                pieces_rotated
            )
            if pieces <= 0:
                weight_per_part = float("inf")
            else:
                sheet_weight = (
                    sheet_l
                    * sheet_w
                    * thickness
                    * 7.85
                ) / 1_000_000
                weight_per_part = sheet_weight / pieces
            # Normalize size for display
            display_l, display_w = sorted(
                [sheet_l, sheet_w]
            )
            option = {
                "sheet_size": f"{display_l} × {display_w}",
                "sheet_length": sheet_l,
                "sheet_width": sheet_w,
                "num_parts": pieces,
                "weight_per_part": (
                    0 if weight_per_part == float("inf")
                    else round(weight_per_part, 2)
                ),
            }
            all_options.append(option)
            # Winner = lowest weight per part
            if (
                pieces > 0
                and weight_per_part < best_weight_per_part
            ):
                best_weight_per_part = weight_per_part
                best_option = option
        if best_option is None:
            return {
                "optimal": None,
                "message": "No approved sheet size can fit the part.",
                "all_options": all_options,
            }
        # Sort by LOWEST weight per part first
        all_options.sort(
            key=lambda x: (
                x["weight_per_part"] if x["weight_per_part"] > 0 else 999999
            )
        )
        return {
            "all_options": all_options,
            "best_option": best_option,
            "effective_part_length": round(
                effective_length,
                2
            ),
            "effective_part_width": round(
                effective_width,
                2
            ),
        }

    def check_sheet_optimization(
        self,
        employee_id: str,
        part_number: str,
        includes_cutting_allowance: bool = True,
    ) -> dict[str, Any]:
        """Public entry point: validate the supplier's sheet choice against
        all approved sheet sizes and return the optimization result."""
        session = self._ensure_session(employee_id, part_number)
        data = session["extracted_data"]

        result = self._validate_sheet_optimization(data, includes_cutting_allowance)

        # Early exit if dimensions missing
        if "best_option" not in result:
            session["sheet_optimization"] = result
            self._persist_session(session)
            return result

        best = result["best_option"]
        current_sheet_l = float(data.get("sheet_length") or 0)
        current_sheet_w = float(data.get("sheet_width") or 0)
        display_l, display_w = sorted([
            int(current_sheet_l),
            int(current_sheet_w)
        ])
        current_sheet = f"{display_l} \u00d7 {display_w}"

        # Compare current sheet to optimal (either orientation)
        current_normalized = tuple(
            sorted([
                int(current_sheet_l),
                int(current_sheet_w)
            ])
        )
        best_normalized = tuple(
            sorted([
                int(best["sheet_length"]),
                int(best["sheet_width"])
            ])
        )
        is_optimal = (
            current_normalized
            == best_normalized
        )
        result["current_sheet"] = current_sheet
        result["is_optimal"] = is_optimal
        result["includes_cutting_allowance"] = includes_cutting_allowance

        if not is_optimal:
            result["recommendation"] = (
                f"The selected sheet size ({current_sheet}) is not optimal. "
                f"Recommended sheet size: {best['sheet_size']}. "
                f"Please upload the revised costing sheet using the recommended sheet size."
            )

        # Store effective (allowance-adjusted) part dims back into session
        if not includes_cutting_allowance:
            data["effective_part_length"] = result["effective_part_length"]
            data["effective_part_width"] = result["effective_part_width"]

        session["sheet_optimization"] = result
        # Allowance question has been answered
        session["awaiting_allowance_response"] = False
        self._persist_session(session)
        return result

        
    def reject_offer(
        self,
        employee_id: str,
        part_number: str,
        reason: str = "Cost exceeds expected benchmark"
    ):
        session = self._ensure_session(
            employee_id,
            part_number
        )
        session["status"] = "rejected"
        session["negotiation"]["status"] = "rejected"
        session["rejection_remark"] = reason
        session["history"].append(
            {
                "role": "tata",
                "message": f"Offer rejected. Reason: {reason}",
                "timestamp": self._now_iso()
            }
        )
        session["negotiation"]["rounds"].append(
            {
                "action": "reject",
                "reason": reason,
                "timestamp": self._now_iso()
            }
        )
        session["summary"] = (
            f"Quotation rejected by Tata Motors. "
            f"Reason: {reason}"
        )
        self._persist_session(session)
        return self._serialize_session(session)

    def reopen_after_rejection(
        self,
        employee_id: str,
        part_number: str,
    ) -> dict[str, Any]:
        """Re-open a rejected session so the supplier can re-negotiate.

        Resets status to 'active', clears sheet optimization (forces re-validation),
        and carries forward the rejection remark so the supplier can see why.
        """
        session = self._ensure_session(employee_id, part_number)
        if session["status"] != "rejected":
            raise ValueError("Only rejected sessions can be reopened.")

        remark = session.get("rejection_remark", "No reason provided.")
        session["status"] = "active"
        session["negotiation"]["status"] = "active"
        # Force re-validation of sheet optimization
        session["sheet_optimization"] = {}
        session["awaiting_allowance_response"] = False
        session["history"].append(
            {
                "role": "system",
                "message": (
                    f"Session reopened for re-negotiation after rejection. "
                    f"Previous rejection reason: {remark}. "
                    f"Please upload a revised costing sheet to continue."
                ),
                "timestamp": self._now_iso(),
            }
        )
        session["summary"] = self._build_summary(session)
        self._persist_session(session)
        logger.info("Session reopened after rejection: %s:%s", employee_id, part_number)
        return self._serialize_session(session)


    def negotiate_with_supplier(self, extracted_data, supplier_message, history, quote, expected, variance):
        breakdown = self._cost_breakdown(extracted_data)
        prompt = f"""You are a Tata Motors Procurement Negotiation Expert.

    Supplier's quoted total: ₹{quote}
    Our itemised expected cost breakdown (₹): {json.dumps(breakdown, indent=2)}
    Our expected total: ₹{expected}
    Variance: {variance}%

    Negotiation History:
    {json.dumps(history, indent=2)}

    Latest Supplier Message: "{supplier_message}"

    STRICT RULES:
    - Only claim a cost is "included" if its value in the breakdown is > 0.
    If profit/packing/transport are 0, DO NOT say they are included.
    - If the supplier points to a cost we excluded, acknowledge it honestly and
    challenge whether the AMOUNT is justified — never pretend it was counted.
    - Numbers followed by "%" are variance, not a price offer.

        Return JSON ONLY:
    {{ "reply": "", "extracted_offer": null,
       "intent": "offer|question|clarification|correction|agreement|rejection|other" }}

    INTENT DEFINITIONS:
    - "correction": supplier says a value in the costing SHEET is wrong/mistaken
      (e.g. "we entered it wrong", "the sheet has an error").
    - "question": supplier is ASKING us something (e.g. "what is your final offer?").
    - "clarification": supplier is explaining/elaborating, NOT reporting a sheet error.
    - "offer": supplier states a NEW price.
    - "agreement": supplier accepts our counter-offer.
    - "rejection": supplier refuses to reduce further.
    """

        payload = {
            "model": self.groq_model,
            "messages": [
                {"role": "system", "content": "You are a Tata Motors procurement negotiation specialist."},
                {"role": "user", "content": prompt},
            ],
            "temperature": 0.2,
            "response_format": {"type": "json_object"},
        }
        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {self.groq_api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=30,
        )
        response.raise_for_status()
        content = response.json()["choices"][0]["message"]["content"].strip()
        if content.startswith("```"):
            content = content.replace("```json", "").replace("```", "").strip()
        return json.loads(content)

    def _negotiate_heuristic(self, extracted_data: dict, supplier_message: str) -> dict:
        """Rule-based fallback negotiation when no LLM key is configured."""
        quote = float(extracted_data.get("total_cost", 0))
        expected = self._compute_expected_cost(extracted_data)
        variance = 0.0
        if expected > 0:
            variance = round(((quote - expected) / expected) * 100, 2)

        if variance <= 5:
            reply = (
                f"Thank you for your message. Your quoted cost of ₹{quote} is within our "
                f"acceptable range. We are pleased to move forward with these terms."
            )
            counter_offer = quote
            status = "accept"
        elif variance <= 15:
            counter = round(expected * 1.03, 2)
            reply = (
                f"Thank you for your message. Your quoted cost of ₹{quote} is slightly above "
                f"our benchmark. We propose a counter-offer of ₹{counter}. "
                f"Please review and confirm."
            )
            counter_offer = counter
            status = "continue"
        else:
            reply = (
                f"Your quoted cost of ₹{quote} exceeds our benchmark by {variance}%. "
                f"Our expected cost is ₹{expected}. Please revise your costing sheet "
                f"and provide a more competitive offer."
            )
            counter_offer = expected
            status = "continue"

        return {"reply": reply, "counter_offer": counter_offer, "status": status}

    def run_negotiation(self, employee_id, part_number, supplier_message):
        session = self._ensure_session(employee_id, part_number)

        # ── Server-side negotiation gate ──
        if session.get("awaiting_allowance_response"):
            raise ValueError(
                "Please answer the cutting allowance question before negotiating."
            )
        sheet_opt = session.get("sheet_optimization", {})
        if not sheet_opt or sheet_opt.get("is_optimal") is None:
            raise ValueError(
                "Sheet optimization must be validated before negotiation. "
                "Please upload a costing sheet and complete the validation."
            )
        if sheet_opt.get("is_optimal") is False:
            raise ValueError(
                "Sheet size is not optimal. "
                "Please upload the revised costing sheet using the recommended sheet size."
            )
        data = session["extracted_data"]
        msg = supplier_message.lower().strip()
        missing_queries = [
            "missing",
            "missing field",
            "missing fields",
            "details missing",
            "any details missing",
            "what is missing",
            "which fields",
            "required fields"
        ]
        if any(q in msg for q in missing_queries):
            missing = session["missing_fields"]
            if missing:
                reply = (
                    "The following mandatory fields are still missing:\n\n• "
                    + "\n• ".join(missing)
                    + "\n\nPlease provide the missing details via chat or upload a revised costing sheet."
                )
            else:
                reply = (
                    "All mandatory fields have been successfully extracted. "
                    "No additional information is currently required."
                )
            session["history"].append(
                {
                    "role": "assistant",
                    "message": supplier_message,
                    "timestamp": self._now_iso()
                }
            )
            self._persist_session(session)
            return {
                "reply": reply,
                "session": self._serialize_session(session)
            }
        final_offer_queries = [
            "final offer",
            "your offer",
            "best offer",
            "counter offer",
            "counter-offer",
            "what can you offer"
        ]
        if any(q in msg for q in final_offer_queries):
            expected_cost = self._compute_expected_cost(data)
            current_counter = (
                session["negotiation"].get("counter_offer")
                or expected_cost
            )
            reply = (
                f"Based on our cost analysis, our expected cost is "
                f"₹{expected_cost:.2f}. Our current counter-offer is "
                f"₹{current_counter:.2f}. Please confirm if this is acceptable."
            )
            session["history"].append(
                {
                    "role": "supplier",
                    "message": supplier_message,
                    "timestamp": self._now_iso()
                }
            )
            session["history"].append(
                {
                    "role": "assistant",
                    "message": reply,
                    "timestamp": self._now_iso()
                }
            )
            self._persist_session(session)
            return {
                "reply": reply,
                "session": self._serialize_session(session)
            }
        quote = float(data.get("total_cost", 0))
        expected_cost = self._compute_expected_cost(data)

        # ---- STEP 1: LLM leads every turn (understanding + reply + offer) ----
        supplier_offer = None
        llm_reply = None
        intent = "other"
        if self.groq_api_key:
            try:
                baseline_variance = round(((quote - expected_cost) / expected_cost) * 100, 2) if expected_cost > 0 else 0
                llm_out = self.negotiate_with_supplier(
                    data, supplier_message, session["negotiation"]["rounds"],
                    quote, expected_cost, baseline_variance,
                )
                llm_reply = llm_out.get("reply")
                intent = llm_out.get("intent", "other")
                if intent == "correction":
                    reply = (
                        "Thank you for pointing this out. "
                        "We have noted that the costing sheet contains an incorrect value. "
                        "Please upload a revised costing sheet so that we can re-evaluate the quotation."
                    )
                    last = session["history"][-1] if session["history"] else None
                    if (
                        last
                        and last.get("role") == "assistant"
                        and last.get("message") == reply
                    ):
                        return {
                            "reply": reply,
                            "session": self._serialize_session(session)
                        }
                    session["history"].append(
                        {
                            "role": "supplier",
                            "message": supplier_message,
                            "timestamp": self._now_iso()
                        }
                    )
                    session["history"].append(
                        {
                            "role": "assistant",
                            "message": reply,
                            "timestamp": self._now_iso()
                        }
                    )
                    session["negotiation"]["rounds"].append(
                        {
                            "role": "supplier",
                            "message": supplier_message,
                            "intent": intent,
                            "timestamp": self._now_iso()
                        }
                    )
                    session["negotiation"]["rounds"].append(
                        {
                            "role": "buyer_ai",
                            "message": reply,
                            "timestamp": self._now_iso()
                        }
                    )
                    self._persist_session(session)
                    return {
                        "reply": reply,
                        "session": self._serialize_session(session)
                    }
                if intent == "agreement":
                    reply = (
                        "Thank you for accepting the counter-offer. "
                        "The quotation will now be submitted for Tata Motors review."
                    )
                    session["status"] = "submitted_for_review"
                    session["history"].append(
                        {
                            "role": "supplier",
                            "message": supplier_message,
                            "timestamp": self._now_iso()
                        }
                    )
                    session["history"].append(
                        {
                            "role": "assistant",
                            "message": reply,
                            "timestamp": self._now_iso()
                        }
                    )
                    self._persist_session(session)
                    return {
                        "reply": reply,
                        "session": self._serialize_session(session)
                    }
                if intent == "rejection":
                    reply = (
                        "We acknowledge your position. "
                        "However, the quotation remains above our expected benchmark. "
                        "At the current price, the quotation cannot be approved."
                    )
                    session["history"].append(
                        {
                            "role": "supplier",
                            "message": supplier_message,
                            "timestamp": self._now_iso()
                        }
                    )
                    session["history"].append(
                        {
                            "role": "assistant",
                            "message": reply,
                            "timestamp": self._now_iso()
                        }
                    )
                    self._persist_session(session)
                    return {
                        "reply": reply,
                        "session": self._serialize_session(session)
                    }
                raw_offer = llm_out.get("extracted_offer")
                if raw_offer is not None:
                    supplier_offer = float(raw_offer)
            except Exception as e:
                logger.warning("LLM negotiation failed, falling back: %s", str(e))

        # ---- STEP 2: regex only as a cheap sanity check (with % protection) ----
        if supplier_offer is None:
            supplier_offer = self._extract_offer_from_message(supplier_message)

        logger.debug("Negotiation: intent=%s, offer=%s, expected=%s", intent, supplier_offer, expected_cost)

        # ---- STEP 3+4: CODE owns the decision + counter-offer math (auditable) ----
        if supplier_offer is not None:
            variance = round(((supplier_offer - expected_cost) / expected_cost) * 100, 2) if expected_cost > 0 else 0

            if variance <= 5:
                decision_reply = (
                    f"Thank you. Your revised offer of ₹{supplier_offer:.2f} is within our "
                    f"acceptable range. The quotation can now be submitted for approval."
                )
                counter_offer = supplier_offer
                status = "accepted"
                session["status"] = "submitted_for_review"
            elif variance <= 15:
                counter_offer = round(expected_cost * 1.03, 2)
                challenge = self._build_negotiation_question(
                    data,
                    session
                )
                decision_reply = (
                    f"Thank you for revising the offer to ₹{supplier_offer:.2f}. "
                    f"Our counter-offer is ₹{counter_offer:.2f}. "
                    f"{challenge}"
                )
                status = "continue"
            else:
                counter_offer = round(expected_cost * 1.02, 2)
                challenge = self._build_negotiation_question(
                    data,
                    session
                )
                decision_reply = (
                    f"Your offer of ₹{supplier_offer:.2f} remains above our expected cost of "
                    f"₹{expected_cost:.2f} ({variance}% variance). "
                    f"Our counter-offer is ₹{counter_offer:.2f}. "
                    f"{challenge}"
                )
                status = "continue"

            # Prefer the LLM's natural phrasing, but numbers came from code above.
            reply = decision_reply
            result = {"reply": reply, "counter_offer": counter_offer, "status": status}

        else:
            # No genuine price this turn → keep the conversation going with the LLM's reply.
            if llm_reply:
                result = {"reply": llm_reply, "counter_offer": session["negotiation"].get("counter_offer", 0), "status": "continue"}
            else:
                result = self._negotiate_heuristic(data, supplier_message)  # no-key fallback

        # ---- STEP 5: persist rounds + history ----
        session["negotiation"]["status"] = result["status"]
        session["negotiation"]["rounds"].append(
            {"role": "supplier", "message": supplier_message, "intent": intent, "timestamp": self._now_iso()}
        )
        session["negotiation"]["rounds"].append(
            {"role": "buyer_ai", "message": result["reply"], "counter_offer": result["counter_offer"], "timestamp": self._now_iso()}
        )
        session["history"].append({"role": "supplier", "message": supplier_message,"timestamp": self._now_iso()})
        session["history"].append({"role": "assistant", "message": result["reply"],"timestamp": self._now_iso()})
        session["negotiation"]["counter_offer"] = result["counter_offer"]

        self._persist_session(session)
        return {"reply": result["reply"], "session": self._serialize_session(session)}


    def _extract_offer_from_message(self, message: str):
        text = re.sub(r"\d+(?:\.\d+)?\s*%", " ", message.lower())  # strip "75%" etc.
        patterns = [
            r"(?:₹|rs\.?|inr)\s*(\d+(?:\.\d+)?)",
            r"(\d+(?:\.\d+)?)\s*(?:rupees|rs\.?|inr)",
            r"(?:offer|quote|price|cost|at|for)\s*(?:of\s*)?₹?\s*(\d+(?:\.\d+)?)",
        ]
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                try:
                    return float(match.group(1))
                except (ValueError, TypeError):
                    pass
        return None

    COST_FIELDS = [
        "raw_material_cost",
        "conversion_cost",
        "coating_cost",
        "overhead_cost",
        "icc_cost",
        "rejection_cost",
        "profit",
        "packing_cost",
        "transport_cost",
    ]
    NEGOTIABILITY = {
        "raw_material_cost": "low",
        "conversion_cost": "medium",
        "coating_cost": "low",
        "overhead_cost": "medium",
        "icc_cost": "medium",
        "rejection_cost": "medium",
        "profit": "high",
        "packing_cost": "high",
        "transport_cost": "high",
    }
    NEGOTIATION_PRIORITY = [
        "packing_cost",
        "profit",
        "conversion_cost",
        "raw_material_cost",
        "coating_cost",
        "transport_cost",
        "overhead_cost",
        "icc_cost",
        "rejection_cost",
    ]

    def _cost_breakdown(self, data: dict) -> dict:
        return {f: round(float(data.get(f) or 0), 2) for f in self.COST_FIELDS}

    def _rank_negotiation_drivers(self, data: dict):
        breakdown = self._cost_breakdown(data)
        scores = []
        for field, value in breakdown.items():
            if value <= 0:
                continue
            negotiability = self.NEGOTIABILITY.get(field, "low")
            weight = {
                "high": 3,
                "medium": 2,
                "low": 1
            }[negotiability]
            score = value * weight
            scores.append({
                "field": field,
                "value": value,
                "negotiability": negotiability,
                "score": score
            })
        priority_order = {
            field: idx
            for idx, field in enumerate(self.NEGOTIATION_PRIORITY)
        }
        scores.sort(
            key=lambda x: (
                priority_order.get(x["field"], 999),
                -x["value"]
            )
        )
        return scores
    
    def _build_negotiation_question(
        self,
        data: dict,
        session: dict
    ):
        challenged = set(
            session["negotiation"].get(
                "challenged_drivers",
                []
            )
        )
        drivers = [
            d for d in self._rank_negotiation_drivers(data)
            if d["field"] not in challenged
        ]
        drivers = drivers[:1]
        if not drivers:
            return "Could you provide more details regarding your costing?"
        questions = []
        for driver in drivers:
            field = driver["field"]
            value = driver["value"]
            session["negotiation"]["challenged_drivers"].append(
                field
            )
            if field == "packing_cost":
                questions.append(
                    f"Packing cost is ₹{value:.2f} per part. "
                    f"Could you provide details of packaging materials, returnable packing options, "
                    f"and any opportunities for logistics optimization?"
                )
            elif field == "profit":
                questions.append(
                    f"Profit allowance is ₹{value:.2f} per part. "
                    f"Considering potential business volume, is there scope to improve the margin structure?"
                )
            elif field == "raw_material_cost":
                questions.append(
                    f"Raw material cost is ₹{value:.2f} per part. "
                    f"Could you share the material procurement rate, yield assumptions, and scrap percentage used?"
                )
            elif field == "conversion_cost":
                questions.append(
                    f"Conversion cost is ₹{value:.2f} per part. "
                    f"Could you provide a breakdown of the major operations and identify any opportunities for process optimization?"
                )
            elif field == "coating_cost":
                questions.append(
                    f"Coating cost is ₹{value:.2f} per part. "
                    f"Please clarify the coating specification, process, and basis used for arriving at this cost."
                )
            elif field == "transport_cost":
                questions.append(
                    f"Transportation cost is ₹{value:.2f} per part. "
                    f"Can shipment consolidation or alternate logistics methods reduce this cost?"
                )
        return (
            "Our analysis identified the following areas for review:\n\n• "
            + "\n\n• ".join(questions)
        )
    
    def _compute_expected_cost(self, data):
        return round(
            float(data.get("raw_material_cost") or 0)
            + float(data.get("conversion_cost") or 0)
            + float(data.get("coating_cost") or 0)
            + float(data.get("overhead_cost") or 0)
            + float(data.get("icc_cost") or 0)
            + float(data.get("rejection_cost") or 0)
            + float(data.get("packing_cost") or 0)
            + float(data.get("transport_cost") or 0),
            2
        )